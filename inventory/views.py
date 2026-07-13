from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.db.models.functions import Coalesce
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.utils import timezone

import csv
import datetime
import json

from .models import Drug, Batch, Supplier, StockTransaction, Alert, CustomUser, Location, StockTransfer
from .forms import (
    CustomAuthenticationForm, DrugForm, SupplierForm, BatchForm, 
    StockInForm, StockOutForm, CustomUserForm, BulkUploadForm,
    LocationForm, StockTransferForm, StockRequestForm
)
from django.db import transaction
import openpyxl
import io

from .utils import check_and_create_alerts, process_fefo_stock_out

def get_active_location(request):
    """
    Returns the active Location for the current request.
    If the user has a locked location in their profile, that is returned.
    If the user is an admin/has no locked location, we look up their active location in the session.
    Defaults to the Central Store if no session context exists.
    """
    if not request.user.is_authenticated:
        return None
        
    # 1. User has assigned location in profile
    if request.user.location:
        return request.user.location
        
    # 2. Look up in session for admin/unassigned users
    active_loc_id = request.session.get('active_location_id')
    if active_loc_id:
        try:
            return Location.objects.get(id=active_loc_id)
        except Location.DoesNotExist:
            pass
            
    # 3. Default to Central Store (first central store or first location found)
    central = Location.objects.filter(is_central=True).first()
    if not central:
        central = Location.objects.first()
        
    if central:
        request.session['active_location_id'] = central.id
        
    return central


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Welcome back, {user.username}!")
                return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = CustomAuthenticationForm()
    
    return render(request, 'inventory/auth/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.info(request, "Logged out successfully.")
    return redirect('login')


@login_required
def profile_view(request):
    if request.method == 'POST':
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.phone = request.POST.get('phone', user.phone)
        user.save()
        messages.success(request, "Profile updated successfully.")
        return redirect('profile')
    return render(request, 'inventory/auth/profile.html')


@login_required
def dashboard_view(request):
    # Run dynamic checks
    check_and_create_alerts()

    active_location = get_active_location(request)

    # Base Metrics
    total_drugs = Drug.objects.count()
    
    # Active batches (stock remaining > 0 at active location)
    batches = Batch.objects.filter(location=active_location, quantity_remaining__gt=0)
    total_items_in_stock = batches.aggregate(Sum('quantity_remaining'))['quantity_remaining__sum'] or 0
    active_batches_count = batches.count()
    
    # Alerts at active location
    pending_alerts = Alert.objects.filter(location=active_location, status='PENDING')
    active_alerts_count = pending_alerts.count()
    
    critical_alerts = pending_alerts.filter(
        Q(alert_type='EXPIRY', batch__expiry_date__lte=timezone.now().date() + datetime.timedelta(days=30)) | 
        Q(alert_type='REORDER')
    )[:5]

    # Low stock count at active location
    low_stock_drugs = []
    for d in Drug.objects.all():
        stock_at_loc = Batch.objects.filter(drug=d, location=active_location, quantity_remaining__gt=0).aggregate(Sum('quantity_remaining'))['quantity_remaining__sum'] or 0
        if stock_at_loc <= d.reorder_level:
            low_stock_drugs.append(d)
    low_stock_count = len(low_stock_drugs)
    
    # Expired batches count at active location
    expired_batches_count = Batch.objects.filter(location=active_location, expiry_date__lte=timezone.now().date(), quantity_remaining__gt=0).count()

    # Recent Transactions at active location
    recent_transactions = StockTransaction.objects.filter(location=active_location).order_by('-transaction_date')[:5]

    # Category breakdown for Chart / Display
    category_data = Drug.objects.values('category').annotate(
        count=Count('id')
    ).order_by('-count')

    # Expiry zone breakdown at active location
    today = timezone.now().date()
    expiring_soon_count = Batch.objects.filter(
        location=active_location,
        expiry_date__gt=today,
        expiry_date__lte=today + datetime.timedelta(days=90),
        quantity_remaining__gt=0
    ).count()

    # Get list of all locations for context switching
    locations = Location.objects.all().order_by('name')

    context = {
        'total_drugs': total_drugs,
        'total_items_in_stock': total_items_in_stock,
        'active_batches_count': active_batches_count,
        'active_alerts_count': active_alerts_count,
        'low_stock_count': low_stock_count,
        'expired_batches_count': expired_batches_count,
        'expiring_soon_count': expiring_soon_count,
        'critical_alerts': critical_alerts,
        'recent_transactions': recent_transactions,
        'category_data': category_data,
        'active_location': active_location,
        'locations': locations,
    }
    return render(request, 'inventory/dashboard.html', context)


# --- DRUG VIEWS ---
@login_required
def drug_list(request):
    query = request.GET.get('q', '')
    category = request.GET.get('category', '')
    status = request.GET.get('status', '')
    
    active_location = get_active_location(request)
    drugs = Drug.objects.all().order_by('drug_name')
    
    if query:
        drugs = drugs.filter(Q(drug_name__icontains=query) | Q(description__icontains=query))
    if category:
        drugs = drugs.filter(category=category)
        
    # Evaluate stock and filter per active location
    filtered_drugs = []
    for d in drugs:
        stock_at_loc = Batch.objects.filter(drug=d, location=active_location, quantity_remaining__gt=0).aggregate(Sum('quantity_remaining'))['quantity_remaining__sum'] or 0
        d.stock_at_loc = stock_at_loc
        d.is_low_at_loc = stock_at_loc <= d.reorder_level
        
        if status == 'low' and not d.is_low_at_loc:
            continue
        elif status == 'out' and stock_at_loc > 0:
            continue
        elif status == 'normal' and (d.is_low_at_loc or stock_at_loc == 0):
            continue
            
        filtered_drugs.append(d)
        
    categories = Drug.CATEGORY_CHOICES
    
    context = {
        'drugs': filtered_drugs,
        'query': query,
        'selected_category': category,
        'selected_status': status,
        'categories': categories,
        'active_location': active_location,
    }
    return render(request, 'inventory/drugs/list.html', context)


@login_required
def drug_detail(request, pk):
    drug = get_object_or_404(Drug, pk=pk)
    active_location = get_active_location(request)
    batches = drug.batches.filter(location=active_location).order_by('expiry_date')
    alerts = drug.alerts.filter(location=active_location, status='PENDING')
    
    # Calculate stock at active location
    stock_at_loc = batches.filter(quantity_remaining__gt=0).aggregate(Sum('quantity_remaining'))['quantity_remaining__sum'] or 0
    
    # Stock in other locations
    other_stocks = []
    for loc in Location.objects.exclude(id=active_location.id).order_by('name'):
        loc_qty = Batch.objects.filter(drug=drug, location=loc, quantity_remaining__gt=0).aggregate(Sum('quantity_remaining'))['quantity_remaining__sum'] or 0
        other_stocks.append({'location': loc, 'quantity': loc_qty})
        
    context = {
        'drug': drug,
        'batches': batches,
        'alerts': alerts,
        'active_location': active_location,
        'stock_at_loc': stock_at_loc,
        'other_stocks': other_stocks,
    }
    return render(request, 'inventory/drugs/detail.html', context)


@login_required
def drug_create(request):
    if request.method == 'POST':
        form = DrugForm(request.POST)
        if form.is_valid():
            drug = form.save()
            messages.success(request, f"Drug '{drug.drug_name}' added successfully!")
            return redirect('drug_detail', pk=drug.pk)
    else:
        form = DrugForm()
    return render(request, 'inventory/drugs/form.html', {'form': form, 'title': 'Add New Drug'})


@login_required
def drug_edit(request, pk):
    drug = get_object_or_404(Drug, pk=pk)
    if request.method == 'POST':
        form = DrugForm(request.POST, instance=drug)
        if form.is_valid():
            form.save()
            messages.success(request, f"Drug '{drug.drug_name}' updated successfully!")
            return redirect('drug_detail', pk=drug.pk)
    else:
        form = DrugForm(instance=drug)
    return render(request, 'inventory/drugs/form.html', {'form': form, 'title': f'Edit Drug: {drug.drug_name}'})


@login_required
def drug_delete(request, pk):
    drug = get_object_or_404(Drug, pk=pk)
    if request.method == 'POST':
        drug.delete()
        messages.success(request, f"Drug '{drug.drug_name}' deleted successfully.")
        return redirect('drug_list')
    return render(request, 'inventory/drugs/delete_confirm.html', {'object': drug, 'type': 'Drug'})


# --- SUPPLIER VIEWS ---
@login_required
def supplier_list(request):
    query = request.GET.get('q', '')
    suppliers = Supplier.objects.all().order_by('supplier_name')
    if query:
        suppliers = suppliers.filter(
            Q(supplier_name__icontains=query) | 
            Q(contact_person__icontains=query) |
            Q(email__icontains=query)
        )
    return render(request, 'inventory/suppliers/list.html', {'suppliers': suppliers, 'query': query})


@login_required
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f"Supplier '{supplier.supplier_name}' added successfully!")
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'inventory/suppliers/form.html', {'form': form, 'title': 'Add New Supplier'})


@login_required
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, f"Supplier '{supplier.supplier_name}' updated successfully!")
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'inventory/suppliers/form.html', {'form': form, 'title': f'Edit Supplier: {supplier.supplier_name}'})


@login_required
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        messages.success(request, f"Supplier '{supplier.supplier_name}' deleted successfully.")
        return redirect('supplier_list')
    return render(request, 'inventory/drugs/delete_confirm.html', {'object': supplier, 'type': 'Supplier'})


# --- BATCH VIEWS ---
@login_required
def batch_list(request):
    active_location = get_active_location(request)
    query = request.GET.get('q', '')
    batches = Batch.objects.filter(location=active_location).order_by('expiry_date')
    if query:
        batches = batches.filter(
            Q(batch_number__icontains=query) | 
            Q(drug__drug_name__icontains=query)
        )
    return render(request, 'inventory/batches/list.html', {
        'batches': batches,
        'query': query,
        'active_location': active_location,
    })


@login_required
def batch_edit(request, pk):
    active_location = get_active_location(request)
    batch = get_object_or_404(Batch, pk=pk, location=active_location)
    if request.method == 'POST':
        form = BatchForm(request.POST, instance=batch)
        if form.is_valid():
            form.save()
            check_and_create_alerts()
            messages.success(request, f"Batch {batch.batch_number} updated successfully!")
            return redirect('batch_list')
    else:
        form = BatchForm(instance=batch)
    return render(request, 'inventory/batches/form.html', {'form': form, 'title': f'Edit Batch: {batch.batch_number}'})


# --- TRANSACTION & STOCK VIEWS ---
@login_required
def transaction_list(request):
    active_location = get_active_location(request)
    query = request.GET.get('q', '')
    t_type = request.GET.get('type', '')
    
    transactions = StockTransaction.objects.filter(location=active_location).order_by('-transaction_date')
    
    if query:
        transactions = transactions.filter(
            Q(batch__drug__drug_name__icontains=query) | 
            Q(batch__batch_number__icontains=query) |
            Q(reference__icontains=query)
        )
    if t_type:
        transactions = transactions.filter(transaction_type=t_type)
        
    return render(request, 'inventory/transactions/list.html', {
        'transactions': transactions, 
        'query': query,
        'selected_type': t_type,
        'active_location': active_location,
    })


@login_required
def stock_in_view(request):
    active_location = get_active_location(request)
    if request.method == 'POST':
        form = StockInForm(request.POST)
        if form.is_valid():
            drug = form.cleaned_data['drug']
            batch_num = form.cleaned_data['batch_number']
            mfg_date = form.cleaned_data['manufacturing_date']
            exp_date = form.cleaned_data['expiry_date']
            qty = form.cleaned_data['quantity_received']
            supplier = form.cleaned_data['supplier']
            ref = form.cleaned_data['reference']
            
            # Check if this batch already exists for this drug at active location
            batch, created = Batch.objects.get_or_create(
                drug=drug,
                batch_number=batch_num,
                location=active_location,
                defaults={
                    'manufacturing_date': mfg_date,
                    'expiry_date': exp_date,
                    'quantity_received': qty,
                    'quantity_remaining': qty,
                    'supplier': supplier
                }
            )
            
            if not created:
                # Augment stock on existing batch
                batch.quantity_received += qty
                batch.quantity_remaining += qty
                batch.save()
            
            # Record transaction
            StockTransaction.objects.create(
                batch=batch,
                transaction_type='IN',
                quantity=qty,
                user=request.user,
                location=active_location,
                reference=ref or f"Manual Ingestion ({'Created' if created else 'Updated'} Batch)"
            )
            
            check_and_create_alerts()
            messages.success(request, f"Stock-In successful. Added {qty} units of {drug.drug_name} to Batch {batch_num} at {active_location.name}.")
            return redirect('transaction_list')
    else:
        form = StockInForm()
    return render(request, 'inventory/transactions/stock_in.html', {'form': form, 'active_location': active_location})


@login_required
def stock_out_view(request):
    active_location = get_active_location(request)
    if request.method == 'POST':
        form = StockOutForm(request.POST, location=active_location)
        form.fields['drug'].queryset = Drug.objects.annotate(
            stock_at_loc=Coalesce(
                Sum('batches__quantity_remaining', filter=Q(batches__location=active_location)),
                0
            )
        ).order_by('drug_name')
        if form.is_valid():
            method = form.cleaned_data['method']
            qty = form.cleaned_data['quantity']
            ref = form.cleaned_data['reference']
            
            try:
                if method == 'FEFO':
                    drug = form.cleaned_data['drug']
                    process_fefo_stock_out(drug, qty, request.user, active_location, ref)
                    messages.success(request, f"Successfully distributed {qty} units of {drug.drug_name} from {active_location.name} using FEFO.")
                else:
                    batch = form.cleaned_data['batch']
                    # Manual Batch selection
                    batch.quantity_remaining -= qty
                    batch.save()
                    
                    StockTransaction.objects.create(
                        batch=batch,
                        transaction_type='OUT',
                        quantity=qty,
                        user=request.user,
                        location=active_location,
                        reference=ref or f"Manual Batch Dispatch"
                    )
                    check_and_create_alerts()
                    messages.success(request, f"Successfully distributed {qty} units of {batch.drug.drug_name} from Batch {batch.batch_number} at {active_location.name}.")
                
                return redirect('transaction_list')
            except ValidationError as e:
                form.add_error(None, e.message)
    else:
        # Check if pre-populated from a drug page
        drug_id = request.GET.get('drug_id')
        initial = {'method': 'FEFO'}
        if drug_id:
            drug = get_object_or_404(Drug, id=drug_id)
            initial['drug'] = drug
        form = StockOutForm(initial=initial, location=active_location)
        form.fields['drug'].queryset = Drug.objects.annotate(
            stock_at_loc=Coalesce(
                Sum('batches__quantity_remaining', filter=Q(batches__location=active_location)),
                0
            )
        ).order_by('drug_name')
        
    return render(request, 'inventory/transactions/stock_out.html', {'form': form, 'active_location': active_location})


# --- ALERTS VIEWS ---
@login_required
def alerts_list(request):
    check_and_create_alerts()
    active_location = get_active_location(request)
    a_type = request.GET.get('type', '')
    status = request.GET.get('status', 'PENDING') # Default to pending alerts
    
    alerts = Alert.objects.filter(location=active_location).order_by('-alert_date')
    
    if a_type:
        alerts = alerts.filter(alert_type=a_type)
    if status:
        alerts = alerts.filter(status=status)
        
    return render(request, 'inventory/alerts/list.html', {
        'alerts': alerts,
        'selected_type': a_type,
        'selected_status': status,
        'active_location': active_location,
    })


@login_required
def resolve_alert(request, pk):
    active_location = get_active_location(request)
    alert = get_object_or_404(Alert, pk=pk, location=active_location)
    alert.status = 'RESOLVED'
    alert.message = f"RESOLVED manually by {request.user.username}: " + alert.message
    alert.save()
    messages.success(request, "Alert marked as resolved.")
    return redirect('alerts_list')


# --- REPORTS & EXPORT VIEWS ---
def _get_expiring_batches_queryset(request, today):
    expiry_days = request.GET.get('expiry_days', '180')
    expiry_start = request.GET.get('expiry_start', '')
    expiry_end = request.GET.get('expiry_end', '')

    active_location = get_active_location(request)

    filters = {
        'quantity_remaining__gt': 0,
        'location': active_location,
    }

    if expiry_days == 'custom':
        if expiry_start:
            try:
                filters['expiry_date__gte'] = datetime.datetime.strptime(expiry_start, '%Y-%m-%d').date()
            except ValueError:
                pass
        if expiry_end:
            try:
                filters['expiry_date__lte'] = datetime.datetime.strptime(expiry_end, '%Y-%m-%d').date()
            except ValueError:
                pass
    elif expiry_days == 'all':
        # Show all future expirations
        filters['expiry_date__gt'] = today
    else:
        # Defaults or preset days
        try:
            days = int(expiry_days)
        except ValueError:
            days = 180
            expiry_days = '180'
        
        filters['expiry_date__gt'] = today
        filters['expiry_date__lte'] = today + datetime.timedelta(days=days)

    return Batch.objects.filter(**filters).order_by('expiry_date'), expiry_days, expiry_start, expiry_end


@login_required
def reports_dashboard(request):
    # Dynamic calculations
    today = timezone.now().date()
    active_location = get_active_location(request)
    
    # 1. Slow moving vs fast moving at this location
    # We can evaluate drugs by total quantity sold/distributed in the last 30 days
    last_30_days = timezone.now() - datetime.timedelta(days=30)
    drug_sales = StockTransaction.objects.filter(
        location=active_location,
        transaction_type='OUT',
        transaction_date__gte=last_30_days
    ).values('batch__drug__drug_name', 'batch__drug__category').annotate(
        total_dispensed=Sum('quantity')
    ).order_by('-total_dispensed')[:10]

    # 2. Expiry forecasts (with filters)
    expiring_batches, expiry_days, expiry_start, expiry_end = _get_expiring_batches_queryset(request, today)

    # 3. Overall Inventory Value approximation (if cost were tracked, but since we don't have it, let's display counts)
    category_summary = Drug.objects.values('category').annotate(
        total_types=Count('id')
    ).order_by('-total_types')

    context = {
        'drug_sales': drug_sales,
        'expiring_batches': expiring_batches,
        'category_summary': category_summary,
        'expiry_days': expiry_days,
        'expiry_start': expiry_start,
        'expiry_end': expiry_end,
        'active_location': active_location,
    }
    return render(request, 'inventory/reports/dashboard.html', context)


@login_required
def export_expiring_batches_csv(request):
    today = timezone.now().date()
    expiring_batches, expiry_days, expiry_start, expiry_end = _get_expiring_batches_queryset(request, today)
    active_location = get_active_location(request)
    
    response = HttpResponse(content_type='text/csv')
    
    # Generate custom filename indicating filters
    if expiry_days == 'custom':
        filename = f"expiring_batches_{active_location.name.replace(' ', '_')}_custom_{expiry_start}_to_{expiry_end}.csv"
    elif expiry_days == 'all':
        filename = f"expiring_batches_{active_location.name.replace(' ', '_')}_all_future.csv"
    else:
        filename = f"expiring_batches_{active_location.name.replace(' ', '_')}_{expiry_days}_days.csv"
        
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(['Batch Number', 'Drug Name', 'Category', 'Expiry Date', 'Days Remaining', 'Quantity Remaining', 'Expiry Status Zone'])
    
    for batch in expiring_batches:
        writer.writerow([
            batch.batch_number,
            batch.drug.drug_name,
            batch.drug.category,
            batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else '',
            batch.days_until_expiry,
            batch.quantity_remaining,
            batch.expiry_status_zone
        ])
        
    return response


@login_required
def export_drugs_csv(request):
    active_location = get_active_location(request)
    response = HttpResponse(content_type='text/csv')
    filename = f"drugs_inventory_report_{active_location.name.replace(' ', '_')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(['Drug Name', 'Category', 'Unit', 'Location Stock', 'Min Stock', 'Reorder Level', 'Max Stock', 'Status'])
    
    drugs = Drug.objects.all().order_by('drug_name')
    for d in drugs:
        stock_at_loc = Batch.objects.filter(drug=d, location=active_location, quantity_remaining__gt=0).aggregate(Sum('quantity_remaining'))['quantity_remaining__sum'] or 0
        status = 'Normal'
        if stock_at_loc == 0:
            status = 'Out of Stock'
        elif stock_at_loc <= d.reorder_level:
            status = 'Low Stock'
            
        writer.writerow([
            d.drug_name, 
            d.category, 
            d.unit, 
            stock_at_loc, 
            d.min_stock, 
            d.reorder_level, 
            d.max_stock, 
            status
        ])
        
    return response


@login_required
def export_transactions_csv(request):
    active_location = get_active_location(request)
    response = HttpResponse(content_type='text/csv')
    filename = f"stock_transactions_log_{active_location.name.replace(' ', '_')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(['Transaction ID', 'Date', 'Type', 'Drug Name', 'Batch Number', 'Quantity', 'User', 'Reference'])
    
    transactions = StockTransaction.objects.filter(location=active_location).order_by('-transaction_date')
    for t in transactions:
        writer.writerow([
            t.id,
            t.transaction_date.strftime('%Y-%m-%d %H:%M:%S'),
            t.get_transaction_type_display(),
            t.batch.drug.drug_name,
            t.batch.batch_number,
            t.quantity,
            t.user.username,
            t.reference or ''
        ])
        
    return response


# --- USER MANAGEMENT VIEWS (ADMIN ONLY) ---

def admin_required(view_func):
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.role != 'ADMIN':
            messages.error(request, "Access Denied: Admin authorization required.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper

@admin_required
def user_list(request):
    query = request.GET.get('q', '')
    users = CustomUser.objects.all().order_by('username')
    if query:
        users = users.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    return render(request, 'inventory/users/list.html', {'users': users, 'query': query})

@admin_required
def user_create(request):
    if request.method == 'POST':
        form = CustomUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"User '{user.username}' created successfully!")
            return redirect('user_list')
    else:
        form = CustomUserForm()
    return render(request, 'inventory/users/form.html', {'form': form, 'title': 'Create New User'})

@admin_required
def user_edit(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        form = CustomUserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f"User '{user.username}' updated successfully!")
            return redirect('user_list')
    else:
        form = CustomUserForm(instance=user)
    return render(request, 'inventory/users/form.html', {'form': form, 'title': f'Edit User: {user.username}'})

@admin_required
def user_delete(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if user == request.user:
        messages.error(request, "Safety block: You cannot delete your own logged-in administrator account.")
        return redirect('user_list')
    if request.method == 'POST':
        user.delete()
        messages.success(request, f"User '{user.username}' deleted successfully.")
        return redirect('user_list')
    return render(request, 'inventory/drugs/delete_confirm.html', {'object': user, 'type': 'User'})


@login_required
def download_drug_template(request):
    """Generates a downloadable CSV template for bulk drug uploads."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="drug_upload_template.csv"'
    
    writer = csv.writer(response)
    # Headers
    writer.writerow(['drug_name', 'category', 'unit', 'reorder_level', 'min_stock', 'max_stock', 'description'])
    # Sample rows
    writer.writerow(['Paracetamol 500mg', 'TABLET', 'Tablet', '100', '50', '1000', 'For pain relief and fever reduction.'])
    writer.writerow(['Amoxicillin 250mg', 'CAPSULE', 'Capsule', '50', '20', '500', 'Broad-spectrum antibiotic.'])
    writer.writerow(['Cough Syrup', 'SYRUP', 'Bottle', '30', '10', '200', 'Soothing cough formula.'])
    
    return response


@login_required
def bulk_upload_drugs(request):
    errors = []
    success_count = 0
    skipped_count = 0
    
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            name = uploaded_file.name.lower()
            
            rows_data = []
            headers = []
            
            try:
                if name.endswith('.csv'):
                    file_data = uploaded_file.read().decode('utf-8-sig')
                    io_string = io.StringIO(file_data)
                    reader = csv.reader(io_string)
                    rows = list(reader)
                    if rows:
                        headers = [h.strip().lower() for h in rows[0]]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(cell.strip() for cell in r):
                                continue # skip blank rows
                            row_dict = {}
                            for i, h in enumerate(headers):
                                row_dict[h] = r[i].strip() if i < len(r) else ""
                            row_dict['_row_num'] = r_idx
                            rows_data.append(row_dict)
                else: # .xlsx
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheet = wb.active
                    rows = list(sheet.iter_rows(values_only=True))
                    if rows:
                        headers = [str(h).strip().lower() for h in rows[0] if h is not None]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(cell is not None and str(cell).strip() for cell in r):
                                continue # skip blank rows
                            row_dict = {}
                            for i, h in enumerate(headers):
                                val = r[i] if i < len(r) else ""
                                row_dict[h] = str(val).strip() if val is not None else ""
                            row_dict['_row_num'] = r_idx
                            rows_data.append(row_dict)
            except Exception as e:
                errors.append(f"Failed to parse file: {str(e)}")
                
            # Check required headers
            required_fields = ['drug_name', 'category', 'unit', 'reorder_level', 'min_stock', 'max_stock']
            missing_fields = [f for f in required_fields if f not in headers]
            if missing_fields:
                errors.append(f"Missing required columns in header: {', '.join(missing_fields)}")
                
            if not errors and not rows_data:
                errors.append("The uploaded file does not contain any data rows.")

            if not errors:
                # Validate and insert in transaction
                valid_categories = {k.lower(): k for k, v in Drug.CATEGORY_CHOICES}
                valid_units = {k.lower(): k for k, v in Drug.UNIT_CHOICES}
                
                try:
                    with transaction.atomic():
                        drugs_to_create = []
                        for row in rows_data:
                            row_num = row['_row_num']
                            d_name = row.get('drug_name', '')
                            cat_raw = row.get('category', '').strip().lower()
                            unit_raw = row.get('unit', '').strip().lower()
                            r_lvl = row.get('reorder_level', '')
                            min_s = row.get('min_stock', '')
                            max_s = row.get('max_stock', '')
                            desc = row.get('description', '')
                            
                            # Validations
                            row_errors = []
                            if not d_name:
                                row_errors.append("Drug name is required.")
                                
                            cat = ""
                            if not cat_raw:
                                row_errors.append("Category is required.")
                            elif cat_raw not in valid_categories:
                                row_errors.append(f"Invalid category '{row.get('category', '')}'. Must be one of: {', '.join(valid_categories.values())}.")
                            else:
                                cat = valid_categories[cat_raw]
                                
                            unit = ""
                            if not unit_raw:
                                row_errors.append("Unit is required.")
                            elif unit_raw not in valid_units:
                                row_errors.append(f"Invalid unit '{row.get('unit', '')}'. Must be one of: {', '.join(valid_units.values())}.")
                            else:
                                unit = valid_units[unit_raw]
                                
                            # Numeric validations
                            try:
                                r_lvl_val = int(r_lvl)
                                if r_lvl_val < 0:
                                    row_errors.append("Reorder level must be >= 0.")
                            except ValueError:
                                row_errors.append(f"Invalid reorder level '{r_lvl}'. Must be an integer.")
                                
                            try:
                                min_s_val = int(min_s)
                                if min_s_val < 0:
                                    row_errors.append("Min stock must be >= 0.")
                            except ValueError:
                                row_errors.append(f"Invalid min stock '{min_s}'. Must be an integer.")
                                
                            try:
                                max_s_val = int(max_s)
                                if max_s_val < 0:
                                    row_errors.append("Max stock must be >= 0.")
                            except ValueError:
                                row_errors.append(f"Invalid max stock '{max_s}'. Must be an integer.")
                            
                            if not row_errors:
                                # Check duplicate in DB or list
                                if Drug.objects.filter(drug_name__iexact=d_name).exists() or any(d.drug_name.lower() == d_name.lower() for d in drugs_to_create):
                                    skipped_count += 1
                                else:
                                    drugs_to_create.append(Drug(
                                        drug_name=d_name,
                                        category=cat,
                                        unit=unit,
                                        reorder_level=r_lvl_val,
                                        min_stock=min_s_val,
                                        max_stock=max_s_val,
                                        description=desc
                                    ))
                            else:
                                for err in row_errors:
                                    errors.append(f"Row {row_num}: {err}")
                        
                        if not errors:
                            # Save all
                            Drug.objects.bulk_create(drugs_to_create)
                            success_count = len(drugs_to_create)
                        else:
                            # Raise exception to force transaction rollback
                            raise ValidationError("Validation failed.")
                except ValidationError:
                    # Expected if we have errors, transaction is rolled back
                    pass
                except Exception as e:
                    errors.append(f"Database error occurred: {str(e)}")

            if not errors:
                msg = f"Successfully imported {success_count} new drugs."
                if skipped_count > 0:
                    msg += f" {skipped_count} existing drugs were skipped."
                messages.success(request, msg)
                return redirect('drug_list')
    else:
        form = BulkUploadForm()
        
    return render(request, 'inventory/drugs/bulk_upload.html', {
        'form': form,
        'errors': errors,
        'title': 'Bulk Upload Drugs'
    })


@login_required
def download_stock_in_template(request):
    """Generates a downloadable CSV template for bulk Stock-In uploads."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="stock_in_upload_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['drug_name', 'batch_number', 'manufacturing_date', 'expiry_date', 'quantity_received', 'supplier_name', 'reference'])
    
    # Try to write matching real catalog items or fallback to Paracetamol
    drug_sample = Drug.objects.first()
    drug_name = drug_sample.drug_name if drug_sample else "Paracetamol 500mg"
    supplier_sample = Supplier.objects.first()
    supplier_name = supplier_sample.supplier_name if supplier_sample else "Test Pharma Inc"
    
    today = datetime.date.today()
    mfg_date = (today - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
    exp_date = (today + datetime.timedelta(days=365)).strftime('%Y-%m-%d')
    
    writer.writerow([drug_name, 'BATCH-001', mfg_date, exp_date, '500', supplier_name, 'INV-2026-001'])
    writer.writerow([drug_name, 'BATCH-002', mfg_date, exp_date, '250', supplier_name, 'PO-99238'])
    
    return response


@login_required
def bulk_stock_in(request):
    errors = []
    success_count = 0
    
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            name = uploaded_file.name.lower()
            
            rows_data = []
            headers = []
            
            try:
                if name.endswith('.csv'):
                    file_data = uploaded_file.read().decode('utf-8-sig')
                    io_string = io.StringIO(file_data)
                    reader = csv.reader(io_string)
                    rows = list(reader)
                    if rows:
                        headers = [h.strip().lower() for h in rows[0]]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(cell.strip() for cell in r):
                                continue # skip blank rows
                            row_dict = {}
                            for i, h in enumerate(headers):
                                row_dict[h] = r[i].strip() if i < len(r) else ""
                            row_dict['_row_num'] = r_idx
                            rows_data.append(row_dict)
                else: # .xlsx
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheet = wb.active
                    rows = list(sheet.iter_rows(values_only=True))
                    if rows:
                        headers = [str(h).strip().lower() for h in rows[0] if h is not None]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(cell is not None and str(cell).strip() for cell in r):
                                continue # skip blank rows
                            row_dict = {}
                            for i, h in enumerate(headers):
                                val = r[i] if i < len(r) else ""
                                row_dict[h] = str(val).strip() if val is not None else ""
                            row_dict['_row_num'] = r_idx
                            rows_data.append(row_dict)
            except Exception as e:
                errors.append(f"Failed to parse file: {str(e)}")
                
            # Check required headers
            required_fields = ['drug_name', 'batch_number', 'manufacturing_date', 'expiry_date', 'quantity_received']
            missing_fields = [f for f in required_fields if f not in headers]
            if missing_fields:
                errors.append(f"Missing required columns in header: {', '.join(missing_fields)}")
                
            if not errors and not rows_data:
                errors.append("The uploaded file does not contain any data rows.")

            if not errors:
                # --- Pass 1: validate all rows, build cleaned payload ---
                validated_rows = []
                for row in rows_data:
                    row_num = row['_row_num']
                    d_name = row.get('drug_name', '').strip()
                    b_num = row.get('batch_number', '').strip()
                    mfg_str = row.get('manufacturing_date', '').strip()
                    exp_str = row.get('expiry_date', '').strip()
                    qty_str = row.get('quantity_received', '').strip()
                    sup_name = row.get('supplier_name', '').strip()
                    ref = row.get('reference', '').strip()

                    row_errors = []

                    # Validate Drug
                    drug_obj = None
                    if not d_name:
                        row_errors.append("Drug name is required.")
                    else:
                        try:
                            drug_obj = Drug.objects.get(drug_name__iexact=d_name)
                        except Drug.DoesNotExist:
                            row_errors.append(f"Drug '{d_name}' does not exist in catalog. Please add it first.")

                    # Validate Batch Number
                    if not b_num:
                        row_errors.append("Batch number is required.")

                    # Validate Manufacturing Date
                    mfg_date = None
                    if not mfg_str:
                        row_errors.append("Manufacturing date is required.")
                    else:
                        try:
                            mfg_date = datetime.datetime.strptime(mfg_str.split(' ')[0], '%Y-%m-%d').date()
                        except ValueError:
                            row_errors.append(f"Invalid manufacturing date format '{mfg_str}'. Use YYYY-MM-DD.")

                    # Validate Expiry Date
                    exp_date = None
                    if not exp_str:
                        row_errors.append("Expiry date is required.")
                    else:
                        try:
                            exp_date = datetime.datetime.strptime(exp_str.split(' ')[0], '%Y-%m-%d').date()
                        except ValueError:
                            row_errors.append(f"Invalid expiry date format '{exp_str}'. Use YYYY-MM-DD.")

                    # Date ordering
                    if mfg_date and exp_date and mfg_date >= exp_date:
                        row_errors.append("Expiry date must be after manufacturing date.")

                    # Validate Quantity
                    qty_val = 0
                    if not qty_str:
                        row_errors.append("Quantity received is required.")
                    else:
                        try:
                            qty_val = int(qty_str)
                            if qty_val < 1:
                                row_errors.append("Quantity received must be at least 1.")
                        except ValueError:
                            row_errors.append(f"Invalid quantity '{qty_str}'. Must be an integer.")

                    if row_errors:
                        for err in row_errors:
                            errors.append(f"Row {row_num}: {err}")
                    else:
                        validated_rows.append({
                            'drug': drug_obj,
                            'batch_number': b_num,
                            'mfg_date': mfg_date,
                            'exp_date': exp_date,
                            'qty': qty_val,
                            'supplier_name': sup_name,
                            'reference': ref,
                        })

                # --- Pass 2: commit only if zero validation errors ---
                if not errors:
                    try:
                        with transaction.atomic():
                            for payload in validated_rows:
                                # Resolve or create supplier
                                supplier_obj = None
                                if payload['supplier_name']:
                                    supplier_obj, _ = Supplier.objects.get_or_create(
                                        supplier_name__iexact=payload['supplier_name'],
                                        defaults={'supplier_name': payload['supplier_name']}
                                    )

                                # Create or augment batch
                                batch, created = Batch.objects.get_or_create(
                                    drug=payload['drug'],
                                    batch_number=payload['batch_number'],
                                    defaults={
                                        'manufacturing_date': payload['mfg_date'],
                                        'expiry_date': payload['exp_date'],
                                        'quantity_received': payload['qty'],
                                        'quantity_remaining': payload['qty'],
                                        'supplier': supplier_obj
                                    }
                                )
                                if not created:
                                    batch.quantity_received += payload['qty']
                                    batch.quantity_remaining += payload['qty']
                                    batch.save()

                                # Record Stock Transaction
                                StockTransaction.objects.create(
                                    batch=batch,
                                    transaction_type='IN',
                                    quantity=payload['qty'],
                                    user=request.user,
                                    reference=payload['reference'] or f"Bulk Ingestion ({'Created' if created else 'Updated'} Batch)"
                                )
                                success_count += 1
                    except Exception as e:
                        errors.append(f"Database error occurred: {str(e)}")


            if not errors:
                messages.success(request, f"Successfully ingested {success_count} stock batches via bulk upload.")
                check_and_create_alerts()
                return redirect('transaction_list')
    else:
        form = BulkUploadForm()
        
    return render(request, 'inventory/transactions/bulk_stock_in.html', {
        'form': form,
        'errors': errors,
        'title': 'Bulk Stock-In Ingestion',
        'required_cols': [
            ('drug_name', 'must match an existing catalog drug'),
            ('batch_number', 'unique code for this delivery batch'),
            ('manufacturing_date', 'format: YYYY-MM-DD'),
            ('expiry_date', 'format: YYYY-MM-DD, must be after manufacturing'),
            ('quantity_received', 'whole integer ≥ 1'),
            ('supplier_name', 'optional — auto-created if not found'),
            ('reference', 'optional invoice or PO number'),
        ]
    })


# --- LOCATION VIEWS (ADMIN ONLY) ---
@admin_required
def location_list(request):
    active_location = get_active_location(request)
    locations = Location.objects.all().order_by('name')
    for loc in locations:
        loc.batches_count = Batch.objects.filter(location=loc, quantity_remaining__gt=0).count()
        loc.total_stock = Batch.objects.filter(location=loc, quantity_remaining__gt=0).aggregate(Sum('quantity_remaining'))['quantity_remaining__sum'] or 0
        loc.alerts_count = Alert.objects.filter(location=loc, status='PENDING').count()
    return render(request, 'inventory/locations/list.html', {
        'locations': locations,
        'active_location': active_location
    })

@admin_required
def location_create(request):
    active_location = get_active_location(request)
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            loc = form.save()
            messages.success(request, f"Location '{loc.name}' created successfully!")
            return redirect('location_list')
    else:
        form = LocationForm()
    return render(request, 'inventory/locations/form.html', {
        'form': form,
        'title': 'Create Location',
        'active_location': active_location
    })

@admin_required
def location_edit(request, pk):
    active_location = get_active_location(request)
    loc = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=loc)
        if form.is_valid():
            form.save()
            messages.success(request, f"Location '{loc.name}' updated successfully!")
            return redirect('location_list')
    else:
        form = LocationForm(instance=loc)
    return render(request, 'inventory/locations/form.html', {
        'form': form,
        'title': f'Edit Location: {loc.name}',
        'active_location': active_location
    })

@admin_required
def location_delete(request, pk):
    active_location = get_active_location(request)
    loc = get_object_or_404(Location, pk=pk)
    if loc.is_central:
        messages.error(request, "Cannot delete the Central Store location.")
        return redirect('location_list')
    if request.method == 'POST':
        loc.delete()
        messages.success(request, f"Location '{loc.name}' deleted successfully.")
        return redirect('location_list')
    return render(request, 'inventory/drugs/delete_confirm.html', {
        'object': loc,
        'type': 'Location',
        'active_location': active_location
    })

@login_required
def switch_location(request, pk):
    """Updates the session's active_location_id for the user (Admins only)."""
    if request.user.role != 'ADMIN':
        messages.error(request, "Only administrators are permitted to switch locations.")
        return redirect('dashboard')
        
    if request.user.location:
        messages.error(request, "Your account is locked to a specific location.")
        return redirect('dashboard')
        
    location = get_object_or_404(Location, pk=pk)
    request.session['active_location_id'] = location.id
    messages.success(request, f"Switched active location context to: {location.name}")
    next_url = request.GET.get('next', 'dashboard')
    return redirect(next_url)


# --- STOCK TRANSFERS & REQUESTS ---
@login_required
def transfer_list(request):
    active_location = get_active_location(request)
    
    # 1. Incoming Action Tasks
    incoming_transfers = StockTransfer.objects.filter(
        to_location=active_location,
        status='PENDING_RECEIPT'
    ).order_by('-created_at')
    
    incoming_requests = StockTransfer.objects.filter(
        from_location=active_location,
        transfer_type='REQUEST',
        status='PENDING_APPROVAL'
    ).order_by('-created_at')

    # 2. Outgoing Status Tasks
    outgoing_transfers = StockTransfer.objects.filter(
        from_location=active_location,
        status='PENDING_RECEIPT'
    ).order_by('-created_at')
    
    outgoing_requests = StockTransfer.objects.filter(
        to_location=active_location,
        transfer_type='REQUEST',
        status='PENDING_APPROVAL'
    ).order_by('-created_at')

    # 3. Completed transfers/requests involving this location
    completed_transfers = StockTransfer.objects.filter(
        Q(from_location=active_location) | Q(to_location=active_location)
    ).exclude(
        status__in=['PENDING_RECEIPT', 'PENDING_APPROVAL']
    ).order_by('-updated_at')[:20]

    return render(request, 'inventory/transfers/list.html', {
        'incoming_transfers': incoming_transfers,
        'incoming_requests': incoming_requests,
        'outgoing_transfers': outgoing_transfers,
        'outgoing_requests': outgoing_requests,
        'completed_transfers': completed_transfers,
        'active_location': active_location
    })

@login_required
def transfer_create(request):
    """Direct push transfer of existing stock to another location."""
    active_location = get_active_location(request)
    
    # Query drugs and batches to pass to Alpine.js
    active_batches = Batch.objects.filter(location=active_location, quantity_remaining__gt=0).select_related('drug')
    drugs_with_stock = Drug.objects.filter(batches__in=active_batches).distinct().order_by('drug_name')
    
    drugs_list = [{'id': d.id, 'name': d.drug_name} for d in drugs_with_stock]
    batches_list = [
        {
            'id': b.id,
            'batch_number': b.batch_number,
            'quantity_remaining': b.quantity_remaining,
            'drug_id': b.drug.id,
            'expiry_date': b.expiry_date.strftime('%Y-%m-%d') if b.expiry_date else 'No Expiry'
        }
        for b in active_batches
    ]
    
    initial_items = []
    
    if request.method == 'POST':
        form = StockTransferForm(request.POST, location=active_location)
        items_json = request.POST.get('items_json', '[]')
        try:
            items = json.loads(items_json)
        except json.JSONDecodeError:
            items = []
        
        initial_items = items
        
        if form.is_valid():
            to_location = form.cleaned_data['to_location']
            reference = form.cleaned_data['reference']
            
            errors = []
            if not items:
                errors.append("Please add at least one stock transfer item.")
                
            validated_items = []
            seen_batches = set()
            for idx, item in enumerate(items, start=1):
                drug_id = item.get('drug_id')
                batch_id = item.get('batch_id')
                qty = item.get('quantity')
                
                if not drug_id:
                    errors.append(f"Row {idx}: Please select a drug.")
                    continue
                if not batch_id:
                    errors.append(f"Row {idx}: Please select a batch.")
                    continue
                try:
                    qty = int(qty)
                    if qty <= 0:
                        errors.append(f"Row {idx}: Quantity must be greater than 0.")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {idx}: Invalid quantity.")
                    continue
                    
                # Check for duplicate batch selection
                try:
                    batch_id = int(batch_id)
                except (ValueError, TypeError):
                    errors.append(f"Row {idx}: Invalid batch selection.")
                    continue
                    
                if batch_id in seen_batches:
                    errors.append(f"Row {idx}: Duplicate batch selected. Please combine quantities into a single row.")
                    continue
                seen_batches.add(batch_id)
                
                # Verify batch and quantity
                try:
                    batch = Batch.objects.get(id=batch_id, location=active_location)
                    if batch.quantity_remaining < qty:
                        errors.append(f"Row {idx}: Insufficient stock in Batch {batch.batch_number}. Available: {batch.quantity_remaining}, Requested: {qty}")
                    else:
                        validated_items.append((batch.drug, batch, qty))
                except Batch.DoesNotExist:
                    errors.append(f"Row {idx}: Selected batch does not exist or does not belong to your active location.")
            
            if errors:
                for err in errors:
                    messages.error(request, err)
            else:
                with transaction.atomic():
                    for drug, batch, qty in validated_items:
                        # Deduct stock immediately
                        batch.quantity_remaining -= qty
                        batch.save()
                        
                        # Create transaction log
                        StockTransaction.objects.create(
                            batch=batch,
                            transaction_type='OUT',
                            quantity=qty,
                            user=request.user,
                            location=active_location,
                            reference=f"Transfer in transit to {to_location.name} (Ref: {reference})"
                        )
                        
                        # Create StockTransfer record
                        StockTransfer.objects.create(
                            transfer_type='TRANSFER',
                            status='PENDING_RECEIPT',
                            from_location=active_location,
                            to_location=to_location,
                            drug=drug,
                            batch_number=batch.batch_number,
                            manufacturing_date=batch.manufacturing_date,
                            expiry_date=batch.expiry_date,
                            supplier=batch.supplier,
                            quantity=qty,
                            created_by=request.user,
                            reference=reference
                        )
                check_and_create_alerts()
                messages.success(request, f"Successfully dispatched {len(validated_items)} items to {to_location.name}. Pending acceptance.")
                return redirect('transfer_list')
    else:
        form = StockTransferForm(location=active_location)
        
    return render(request, 'inventory/transfers/form_transfer.html', {
        'form': form,
        'active_location': active_location,
        'title': 'New Push Stock Transfer',
        'drugs_json': json.dumps(drugs_list),
        'batches_json': json.dumps(batches_list),
        'initial_items_json': json.dumps(initial_items)
    })

@login_required
def request_create(request):
    """Pull request of stock from another location."""
    active_location = get_active_location(request)
    
    # Query all drugs in the catalog
    drugs = Drug.objects.all().order_by('drug_name')
    drugs_list = [{'id': d.id, 'name': d.drug_name} for d in drugs]
    
    initial_items = []
    
    # Handle pre-population if requested from drug details page
    from_loc_id = request.GET.get('from_location_id')
    drug_id = request.GET.get('drug_id')
    if drug_id:
        try:
            # Verify drug exists
            drug = Drug.objects.get(id=drug_id)
            initial_items.append({'drug_id': int(drug_id), 'quantity': 1})
        except (Drug.DoesNotExist, ValueError):
            pass
            
    if request.method == 'POST':
        form = StockRequestForm(request.POST, location=active_location)
        items_json = request.POST.get('items_json', '[]')
        try:
            items = json.loads(items_json)
        except json.JSONDecodeError:
            items = []
            
        initial_items = items
        
        if form.is_valid():
            from_location = form.cleaned_data['from_location']
            reference = form.cleaned_data['reference']
            
            errors = []
            if not items:
                errors.append("Please add at least one drug request item.")
                
            validated_items = []
            seen_drugs = set()
            for idx, item in enumerate(items, start=1):
                d_id = item.get('drug_id')
                qty = item.get('quantity')
                
                if not d_id:
                    errors.append(f"Row {idx}: Please select a drug.")
                    continue
                try:
                    qty = int(qty)
                    if qty <= 0:
                        errors.append(f"Row {idx}: Quantity must be greater than 0.")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {idx}: Invalid quantity.")
                    continue
                    
                try:
                    d_id = int(d_id)
                except (ValueError, TypeError):
                    errors.append(f"Row {idx}: Invalid drug selection.")
                    continue
                    
                if d_id in seen_drugs:
                    errors.append(f"Row {idx}: Duplicate drug selected. Please combine quantities into a single row.")
                    continue
                seen_drugs.add(d_id)
                
                try:
                    drug = Drug.objects.get(id=d_id)
                    validated_items.append((drug, qty))
                except Drug.DoesNotExist:
                    errors.append(f"Row {idx}: Selected drug does not exist.")
            
            if errors:
                for err in errors:
                    messages.error(request, err)
            else:
                with transaction.atomic():
                    for drug, qty in validated_items:
                        StockTransfer.objects.create(
                            transfer_type='REQUEST',
                            status='PENDING_APPROVAL',
                            from_location=from_location,
                            to_location=active_location,
                            drug=drug,
                            quantity=qty,
                            created_by=request.user,
                            reference=reference
                        )
                messages.success(request, f"Requested {len(validated_items)} drugs from {from_location.name}. Pending approval.")
                return redirect('transfer_list')
    else:
        initial = {}
        if from_loc_id:
            try:
                initial['from_location'] = get_object_or_404(Location, id=from_loc_id)
            except Exception:
                pass
        form = StockRequestForm(initial=initial, location=active_location)
        
    return render(request, 'inventory/transfers/form_request.html', {
        'form': form,
        'active_location': active_location,
        'title': 'Request Stock (Pull)',
        'drugs_json': json.dumps(drugs_list),
        'initial_items_json': json.dumps(initial_items)
    })

@login_required
def transfer_approve_dispatch(request, pk):
    """Approve a stock request, choosing a batch to dispatch it from."""
    active_location = get_active_location(request)
    transfer = get_object_or_404(StockTransfer, pk=pk, from_location=active_location, status='PENDING_APPROVAL')
    
    if request.method == 'POST':
        batch_id = request.POST.get('batch')
        if not batch_id:
            messages.error(request, "Please select a batch to dispatch.")
        else:
            batch = get_object_or_404(Batch, id=batch_id, location=active_location)
            if batch.quantity_remaining < transfer.quantity:
                messages.error(request, f"Insufficient stock in Batch {batch.batch_number}. Available: {batch.quantity_remaining}")
            else:
                with transaction.atomic():
                    # Deduct stock
                    batch.quantity_remaining -= transfer.quantity
                    batch.save()

                    # Record transaction log
                    StockTransaction.objects.create(
                        batch=batch,
                        transaction_type='OUT',
                        quantity=transfer.quantity,
                        user=request.user,
                        location=active_location,
                        reference=f"Dispatched requested stock to {transfer.to_location.name} (Request ID: {transfer.id})"
                    )

                    # Update transfer with batch details and move to PENDING_RECEIPT (Transit)
                    transfer.batch_number = batch.batch_number
                    transfer.manufacturing_date = batch.manufacturing_date
                    transfer.expiry_date = batch.expiry_date
                    transfer.supplier = batch.supplier
                    transfer.status = 'PENDING_RECEIPT'
                    transfer.save()
                    
                check_and_create_alerts()
                messages.success(request, f"Request approved. Dispatched Batch {batch.batch_number} to {transfer.to_location.name}.")
                return redirect('transfer_list')

    batches = Batch.objects.filter(drug=transfer.drug, location=active_location, quantity_remaining__gt=0).order_by('expiry_date')
    return render(request, 'inventory/transfers/approve_dispatch.html', {
        'transfer': transfer,
        'batches': batches,
        'active_location': active_location
    })

@login_required
def transfer_accept(request, pk):
    """Accept incoming transfer or request into stock."""
    active_location = get_active_location(request)
    transfer = get_object_or_404(StockTransfer, pk=pk, to_location=active_location, status='PENDING_RECEIPT')
    
    with transaction.atomic():
        # Get or create batch in target location
        batch, created = Batch.objects.get_or_create(
            drug=transfer.drug,
            batch_number=transfer.batch_number,
            location=active_location,
            defaults={
                'manufacturing_date': transfer.manufacturing_date,
                'expiry_date': transfer.expiry_date,
                'quantity_received': transfer.quantity,
                'quantity_remaining': transfer.quantity,
                'supplier': transfer.supplier
            }
        )
        if not created:
            batch.quantity_received += transfer.quantity
            batch.quantity_remaining += transfer.quantity
            batch.save()
            
        # Record transaction log
        StockTransaction.objects.create(
            batch=batch,
            transaction_type='IN',
            quantity=transfer.quantity,
            user=request.user,
            location=active_location,
            reference=f"Stock Transfer received from {transfer.from_location.name}"
        )
        
        # Update transfer record
        transfer.status = 'ACCEPTED'
        transfer.save()

    check_and_create_alerts()
    messages.success(request, f"Successfully received {transfer.quantity} units of {transfer.drug.drug_name} into Batch {transfer.batch_number} at {active_location.name}.")
    return redirect('transfer_list')

@login_required
def transfer_reject(request, pk):
    """Reject incoming transfer. Stock goes back to the sender."""
    active_location = get_active_location(request)
    transfer = get_object_or_404(StockTransfer, pk=pk, to_location=active_location, status='PENDING_RECEIPT')
    
    rejection_reason = request.POST.get('rejection_reason', 'Rejected by staff.')

    with transaction.atomic():
        # Re-credit sender batch
        try:
            sender_batch = Batch.objects.get(
                drug=transfer.drug,
                batch_number=transfer.batch_number,
                location=transfer.from_location
            )
            sender_batch.quantity_remaining += transfer.quantity
            sender_batch.save()
        except Batch.DoesNotExist:
            sender_batch = Batch.objects.create(
                drug=transfer.drug,
                batch_number=transfer.batch_number,
                location=transfer.from_location,
                manufacturing_date=transfer.manufacturing_date,
                expiry_date=transfer.expiry_date,
                quantity_received=transfer.quantity,
                quantity_remaining=transfer.quantity,
                supplier=transfer.supplier
            )

        # Record re-credit transaction log at sender location
        StockTransaction.objects.create(
            batch=sender_batch,
            transaction_type='IN',
            quantity=transfer.quantity,
            user=request.user,
            location=transfer.from_location,
            reference=f"Re-credited rejected transfer to {active_location.name}. Reason: {rejection_reason}"
        )

        # Update transfer record
        transfer.status = 'REJECTED'
        transfer.rejection_reason = rejection_reason
        transfer.save()

    check_and_create_alerts()
    messages.warning(request, f"Transfer rejected. Stock returned back to {transfer.from_location.name}.")
    return redirect('transfer_list')

@login_required
def transfer_cancel(request, pk):
    """Cancel a pending request before approval."""
    active_location = get_active_location(request)
    transfer = get_object_or_404(StockTransfer, pk=pk, to_location=active_location, status='PENDING_APPROVAL')
    transfer.status = 'CANCELLED'
    transfer.save()
    messages.info(request, "Request cancelled.")
    return redirect('transfer_list')


# --- BULK STOCK-OUT (DISPENSE) & BULK TRANSFER ---
@login_required
def bulk_stock_out(request):
    active_location = get_active_location(request)
    errors = []
    success_count = 0
    
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            name = uploaded_file.name.lower()
            
            rows_data = []
            headers = []
            
            try:
                if name.endswith('.csv'):
                    file_data = uploaded_file.read().decode('utf-8-sig')
                    io_string = io.StringIO(file_data)
                    reader = csv.reader(io_string)
                    rows = list(reader)
                    if rows:
                        headers = [h.strip().lower() for h in rows[0]]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(cell.strip() for cell in r):
                                continue
                            row_dict = {}
                            for i, h in enumerate(headers):
                                row_dict[h] = r[i].strip() if i < len(r) else ""
                            row_dict['_row_num'] = r_idx
                            rows_data.append(row_dict)
                else: # .xlsx
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheet = wb.active
                    rows = list(sheet.iter_rows(values_only=True))
                    if rows:
                        headers = [str(h).strip().lower() for h in rows[0] if h is not None]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(cell is not None and str(cell).strip() for cell in r):
                                continue
                            row_dict = {}
                            for i, h in enumerate(headers):
                                val = r[i] if i < len(r) else ""
                                row_dict[h] = str(val).strip() if val is not None else ""
                            row_dict['_row_num'] = r_idx
                            rows_data.append(row_dict)
            except Exception as e:
                errors.append(f"Failed to parse file: {str(e)}")
                
            required_fields = ['drug_name', 'quantity']
            missing_fields = [f for f in required_fields if f not in headers]
            if missing_fields:
                errors.append(f"Missing required columns: {', '.join(missing_fields)}")
                
            if not errors and not rows_data:
                errors.append("The uploaded file does not contain any data rows.")
                
            if not errors:
                validated_rows = []
                for row in rows_data:
                    row_num = row['_row_num']
                    d_name = row.get('drug_name', '').strip()
                    qty_str = row.get('quantity', '').strip()
                    ref = row.get('reference', '').strip() or "Bulk Dispense"
                    
                    row_errors = []
                    drug_obj = None
                    if not d_name:
                        row_errors.append("Drug name is required.")
                    else:
                        try:
                            drug_obj = Drug.objects.get(drug_name__iexact=d_name)
                        except Drug.DoesNotExist:
                            row_errors.append(f"Drug '{d_name}' does not exist.")
                            
                    qty_val = 0
                    if not qty_str:
                        row_errors.append("Quantity is required.")
                    else:
                        try:
                            qty_val = int(qty_str)
                            if qty_val < 1:
                                row_errors.append("Quantity must be at least 1.")
                        except ValueError:
                            row_errors.append(f"Invalid quantity '{qty_str}'.")
                            
                    # Check stock if drug exists
                    if drug_obj and qty_val > 0:
                        total_available = Batch.objects.filter(drug=drug_obj, location=active_location, quantity_remaining__gt=0).aggregate(Sum('quantity_remaining'))['quantity_remaining__sum'] or 0
                        if total_available < qty_val:
                            row_errors.append(f"Insufficient stock for '{d_name}' at {active_location.name}. Available: {total_available}, Requested: {qty_val}.")
                            
                    if row_errors:
                        for err in row_errors:
                            errors.append(f"Row {row_num}: {err}")
                    else:
                        validated_rows.append({
                            'drug': drug_obj,
                            'qty': qty_val,
                            'reference': ref
                        })
                        
                if not errors:
                    try:
                        with transaction.atomic():
                            for payload in validated_rows:
                                process_fefo_stock_out(payload['drug'], payload['qty'], request.user, active_location, payload['reference'])
                                success_count += 1
                    except Exception as e:
                        errors.append(f"Database error occurred during processing: {str(e)}")
                        
            if not errors:
                messages.success(request, f"Successfully processed bulk dispense for {success_count} drugs.")
                return redirect('transaction_list')
                
    else:
        form = BulkUploadForm()
        
    return render(request, 'inventory/transactions/bulk_stock_out.html', {
        'form': form,
        'errors': errors,
        'active_location': active_location,
        'title': 'Bulk Stock dispensation (FEFO)'
    })

@login_required
def bulk_transfer(request):
    active_location = get_active_location(request)
    errors = []
    success_count = 0
    
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            name = uploaded_file.name.lower()
            
            rows_data = []
            headers = []
            
            try:
                if name.endswith('.csv'):
                    file_data = uploaded_file.read().decode('utf-8-sig')
                    io_string = io.StringIO(file_data)
                    reader = csv.reader(io_string)
                    rows = list(reader)
                    if rows:
                        headers = [h.strip().lower() for h in rows[0]]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(cell.strip() for cell in r):
                                continue
                            row_dict = {}
                            for i, h in enumerate(headers):
                                row_dict[h] = r[i].strip() if i < len(r) else ""
                            row_dict['_row_num'] = r_idx
                            rows_data.append(row_dict)
                else: # .xlsx
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheet = wb.active
                    rows = list(sheet.iter_rows(values_only=True))
                    if rows:
                        headers = [str(h).strip().lower() for h in rows[0] if h is not None]
                        for r_idx, r in enumerate(rows[1:], start=2):
                            if not any(cell is not None and str(cell).strip() for cell in r):
                                continue
                            row_dict = {}
                            for i, h in enumerate(headers):
                                val = r[i] if i < len(r) else ""
                                row_dict[h] = str(val).strip() if val is not None else ""
                            row_dict['_row_num'] = r_idx
                            rows_data.append(row_dict)
            except Exception as e:
                errors.append(f"Failed to parse file: {str(e)}")
                
            required_fields = ['to_location_name', 'drug_name', 'batch_number', 'quantity']
            missing_fields = [f for f in required_fields if f not in headers]
            if missing_fields:
                errors.append(f"Missing required columns: {', '.join(missing_fields)}")
                
            if not errors and not rows_data:
                errors.append("The uploaded file does not contain any data rows.")
                
            if not errors:
                validated_rows = []
                for row in rows_data:
                    row_num = row['_row_num']
                    to_loc_name = row.get('to_location_name', '').strip()
                    d_name = row.get('drug_name', '').strip()
                    b_num = row.get('batch_number', '').strip()
                    qty_str = row.get('quantity', '').strip()
                    ref = row.get('reference', '').strip() or "Bulk Transfer"
                    
                    row_errors = []
                    
                    # Target location validation
                    to_loc_obj = None
                    if not to_loc_name:
                        row_errors.append("Target location name is required.")
                    else:
                        try:
                            to_loc_obj = Location.objects.get(name__iexact=to_loc_name)
                            if to_loc_obj.id == active_location.id:
                                row_errors.append("Cannot transfer to the same location.")
                        except Location.DoesNotExist:
                            row_errors.append(f"Location '{to_loc_name}' does not exist.")
                            
                    # Drug validation
                    drug_obj = None
                    if not d_name:
                        row_errors.append("Drug name is required.")
                    else:
                        try:
                            drug_obj = Drug.objects.get(drug_name__iexact=d_name)
                        except Drug.DoesNotExist:
                            row_errors.append(f"Drug '{d_name}' does not exist.")
                            
                    # Quantity validation
                    qty_val = 0
                    if not qty_str:
                        row_errors.append("Quantity is required.")
                    else:
                        try:
                            qty_val = int(qty_str)
                            if qty_val < 1:
                                row_errors.append("Quantity must be at least 1.")
                        except ValueError:
                            row_errors.append(f"Invalid quantity '{qty_str}'.")
                            
                    # Batch validation at source location
                    batch_obj = None
                    if drug_obj and b_num:
                        try:
                            batch_obj = Batch.objects.get(drug=drug_obj, batch_number=b_num, location=active_location)
                            if qty_val > 0 and batch_obj.quantity_remaining < qty_val:
                                row_errors.append(f"Insufficient stock in Batch '{b_num}'. Available: {batch_obj.quantity_remaining}, Requested: {qty_val}.")
                        except Batch.DoesNotExist:
                            row_errors.append(f"Batch '{b_num}' of '{d_name}' does not exist at {active_location.name}.")
                    elif not b_num:
                        row_errors.append("Batch number is required.")
                        
                    if row_errors:
                        for err in row_errors:
                            errors.append(f"Row {row_num}: {err}")
                    else:
                        validated_rows.append({
                            'to_location': to_loc_obj,
                            'drug': drug_obj,
                            'batch': batch_obj,
                            'qty': qty_val,
                            'reference': ref
                        })
                        
                if not errors:
                    try:
                        with transaction.atomic():
                            for payload in validated_rows:
                                batch = payload['batch']
                                qty = payload['qty']
                                to_loc = payload['to_location']
                                
                                # Deduct source stock
                                batch.quantity_remaining -= qty
                                batch.save()
                                
                                # Record transaction
                                StockTransaction.objects.create(
                                    batch=batch,
                                    transaction_type='OUT',
                                    quantity=qty,
                                    user=request.user,
                                    location=active_location,
                                    reference=f"Bulk transfer dispatch to {to_loc.name} (Ref: {payload['reference']})"
                                )
                                
                                # Create StockTransfer record
                                StockTransfer.objects.create(
                                    transfer_type='TRANSFER',
                                    status='PENDING_RECEIPT',
                                    from_location=active_location,
                                    to_location=to_loc,
                                    drug=payload['drug'],
                                    batch_number=batch.batch_number,
                                    manufacturing_date=batch.manufacturing_date,
                                    expiry_date=batch.expiry_date,
                                    supplier=batch.supplier,
                                    quantity=qty,
                                    created_by=request.user,
                                    reference=payload['reference']
                                )
                                success_count += 1
                    except Exception as e:
                        errors.append(f"Database error occurred: {str(e)}")
                        
            if not errors:
                check_and_create_alerts()
                messages.success(request, f"Successfully uploaded and dispatched {success_count} transfers.")
                return redirect('transfer_list')
                
    else:
        form = BulkUploadForm()
        
    return render(request, 'inventory/transfers/bulk_transfer.html', {
        'form': form,
        'errors': errors,
        'active_location': active_location,
        'title': 'Bulk Stock Transfer Upload'
    })



